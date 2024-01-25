from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import datetime

# Get the current day of the week
current_day = datetime.datetime.now().strftime("%A")

# Load the Excel workbook
workbook = openpyxl.load_workbook(r"C:\Users\USER\Desktop\QUPS_assessment\sample.xlsx")
sheet = workbook[current_day]

# Create a dictionary to map keywords to row numbers
keyword_row_mapping = {f"Keyword{i}": i + 1 for i in range(1, sheet.max_row-1)}
keyword_row_mapping['Keyword10'] = 10  # Add 'Keyword10' mapping

# Configure the Chrome driver
driver = webdriver.Chrome()

# Iterate through each row in the sheet
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    serial_number, keyword = row

    # Open Google Chrome
    driver.get("https://www.google.com")

    # Wait for the search input to be present
    search_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "q"))
    )

    # Type the keyword in the search input
    search_input.send_keys(keyword)

    # Wait for search suggestions to appear
    suggestions = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//ul[@role='listbox']//li"))
    )

    if suggestions:
        # Get the longest and shortest suggestions
        longest_suggestion = max(suggestions, key=lambda x: len(x.text))
        shortest_suggestion = min(suggestions, key=lambda x: len(x.text))

        # Update the Excel sheet
        sheet.cell(row=keyword_row_mapping.get(serial_number, 1), column=3, value=longest_suggestion.text)
        sheet.cell(row=keyword_row_mapping.get(serial_number, 1), column=4, value=shortest_suggestion.text)

        print(f"Processed for Keyword: {keyword}")
        print(f"Longest Suggestion: {longest_suggestion.text}")
        print(f"Shortest Suggestion: {shortest_suggestion.text}")
    else:
        print(f"No suggestions found for Keyword: {keyword}")

    # Clear the search input for the next iteration
    search_input.clear()

# Save the updated workbook
workbook.save(r"C:\Users\USER\Desktop\QUPS_assessment\sample_updated.xlsx")

# Close the Chrome browser
driver.quit()